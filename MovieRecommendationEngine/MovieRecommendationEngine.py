from openpyxl import load_workbook

# Load Excel file
excel_file = '' # Path to file
workbook = load_workbook(excel_file)
sheet = workbook.active

# Locate the beginning of the table
start_row = 31
start_col = 2

# Dictionary containing all ratings
result = {}

# Iterating through rows (for person)
for row in range(start_row, sheet.max_row + 1):
    name = sheet.cell(row, start_col).value

    # Dictionary containing movies with its ratings
    movies = {}

    # Iterating through columns (for movie and rating)
    col = start_col + 1 # Skip name column
    while col <= sheet.max_column:
        movie = sheet.cell(row, col).value
        rating = sheet.cell(row, col + 1).value

        # Add entry only if both movie and rating are present
        if movie and rating:
            movies[movie] = rating

        # Go to the next entry
        col += 2

    # Assigning movies and ratings to person
    result[name] = movies

print(result)
